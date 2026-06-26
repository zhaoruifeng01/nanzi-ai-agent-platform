"""Controlled Excel document tools for configured agents."""
from __future__ import annotations

from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries

from app.core.context import get_current_agent_context
from app.services.ai.tools.document_paths import (
    DocumentPathError,
    resolve_document_input_path,
    resolve_document_output_path,
)
from app.services.ai.tools.generated_file_service import publish
from app.services.ai.tools.tool_compat import tool

_EXTENSIONS = {".xlsx"}
_MAX_PREVIEW_ROWS = 20
_MAX_PREVIEW_COLUMNS = 20
_MAX_RANGE_ROWS = 200
_MAX_RANGE_COLUMNS = 50


def _context():
    context = get_current_agent_context()
    if not context:
        raise DocumentPathError("无法获取当前执行上下文")
    return context


async def _input_path(path: str):
    context = _context()
    return await resolve_document_input_path(
        path,
        allowed_attachment_paths=context.authorized_attachment_paths,
        user_id=context.user_id,
        conversation_id=context.conversation_id,
        allowed_extensions=_EXTENSIONS,
    )


async def _output_path(filename: str):
    context = _context()
    return await resolve_document_output_path(
        filename,
        user_id=context.user_id,
        conversation_id=context.conversation_id,
        allowed_extensions=_EXTENSIONS,
    )


def _artifact_result(output_path, *, summary: str, changes: dict[str, Any]) -> dict[str, Any]:
    artifact = publish(output_path, output_path.name)
    return {
        "status": "ok",
        "summary": summary,
        "changes": changes,
        "artifact": artifact.to_tool_payload(),
    }


@tool
async def excel_document_read(
    action: str,
    path: str,
    sheet_name: str | None = None,
    cell_range: str | None = None,
) -> dict[str, Any]:
    """Inspect an Excel workbook or read a bounded range from an uploaded workbook."""
    if action not in {"inspect", "read_range"}:
        raise DocumentPathError("excel_document_read 仅支持 inspect 或 read_range")
    input_path = await _input_path(path)
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        if action == "inspect":
            sheets = []
            for worksheet in workbook.worksheets:
                preview = [
                    list(row)
                    for row in worksheet.iter_rows(
                        max_row=min(worksheet.max_row, _MAX_PREVIEW_ROWS),
                        max_col=min(worksheet.max_column, _MAX_PREVIEW_COLUMNS),
                        values_only=True,
                    )
                ]
                sheets.append({"name": worksheet.title, "rows": worksheet.max_row, "columns": worksheet.max_column, "preview": preview})
            return {"status": "ok", "summary": f"工作簿包含 {len(sheets)} 个工作表", "data": {"sheets": sheets}, "truncated": False}
        if not sheet_name or not cell_range:
            raise DocumentPathError("read_range 需要 sheet_name 和 cell_range")
        if sheet_name not in workbook.sheetnames:
            raise DocumentPathError("工作表不存在")
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if max_row - min_row + 1 > _MAX_RANGE_ROWS or max_col - min_col + 1 > _MAX_RANGE_COLUMNS:
            raise DocumentPathError("读取范围超过 200 行或 50 列限制")
        worksheet = workbook[sheet_name]
        values = [list(row) for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)]
        return {"status": "ok", "summary": f"已读取 {sheet_name}!{cell_range}", "data": {"values": values}, "truncated": False}
    finally:
        workbook.close()


@tool
async def excel_document_write(
    action: str,
    output_filename: str,
    path: str | None = None,
    sheet_name: str | None = None,
    cells: list[dict[str, Any]] | None = None,
    rows: list[list[Any]] | None = None,
) -> dict[str, Any]:
    """Create or modify an Excel workbook copy and return a download link.

    Copy artifact.download_url verbatim in the final response; never add a protocol or host.
    """
    if action not in {"create", "write_cells", "append_rows", "create_sheet"}:
        raise DocumentPathError("excel_document_write 不支持该操作")
    if action == "create":
        workbook = Workbook()
        worksheet = workbook.active
        if sheet_name:
            worksheet.title = sheet_name
    else:
        if not path:
            raise DocumentPathError("修改工作簿需要 path")
        workbook = load_workbook(await _input_path(path), data_only=False)
        if not sheet_name:
            raise DocumentPathError("修改工作簿需要 sheet_name")
        if action != "create_sheet" and sheet_name not in workbook.sheetnames:
            raise DocumentPathError("工作表不存在")
    changes: dict[str, Any] = {}
    if action == "write_cells":
        worksheet = workbook[sheet_name]
        if not cells:
            raise DocumentPathError("write_cells 需要 cells")
        for cell in cells:
            address = str(cell.get("address") or "")
            if not address:
                raise DocumentPathError("单元格地址不能为空")
            worksheet[address] = cell.get("value")
        changes["written_cells"] = len(cells)
    elif action == "append_rows":
        worksheet = workbook[sheet_name]
        if not rows:
            raise DocumentPathError("append_rows 需要 rows")
        for row in rows:
            worksheet.append(list(row))
        changes["appended_rows"] = len(rows)
    elif action == "create_sheet":
        if sheet_name in workbook.sheetnames:
            raise DocumentPathError("工作表已存在")
        workbook.create_sheet(sheet_name)
        changes["created_sheet"] = sheet_name
    else:
        changes["created_workbook"] = True
    output_path = await _output_path(output_filename)
    workbook.save(output_path)
    workbook.close()
    return _artifact_result(output_path, summary="已生成 Excel 文件", changes=changes)
